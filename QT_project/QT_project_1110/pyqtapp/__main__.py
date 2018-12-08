# -*- coding:utf-8 -*-
import sys
import time
import serial
import functools
import os
import shutil
from openpyxl import Workbook
from pyqtapp.ui_mainwindow import Ui_MainWindow
from pyqtapp.ui_remote_control_dialog import Ui_Dialog
from pyqtapp.ui_waiting_dialog import Waiting_Ui_Dialog
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from .listports import *
from pyqtapp import images_rc


ser = serial.Serial()
ser.baudrate = 115200
ser.timeout=1000


class Waiting(QDialog, Waiting_Ui_Dialog):
    def __init__(self, parent=None):
        super(Waiting, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('Waiting')
        self.movie=QMovie(":image/windmu_hle.gif")
        self.label.setMovie(self.movie)
        self.movie.start()



class RemoteControl(QDialog, Ui_Dialog):    #遙控器對話窗
    YouPress=pyqtSignal(int)
    def __init__(self, parent=None):
        super(RemoteControl, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('Remote Control')
        self.setWindowIcon(QIcon(':/image/icon.ico'))
        self.buttons()

    def buttons(self):
        self.pushButton.clicked.connect(lambda:self.YouPress.emit(int(1)))
        self.pushButton_2.clicked.connect(lambda:self.YouPress.emit(int(2)))
        self.pushButton_3.clicked.connect(lambda:self.YouPress.emit(int(3)))
        self.pushButton_4.clicked.connect(lambda:self.YouPress.emit(int(4)))
        self.pushButton_5.clicked.connect(lambda:self.YouPress.emit(int(5)))
        self.pushButton_6.clicked.connect(lambda:self.YouPress.emit(int(6)))
        self.pushButton_7.clicked.connect(lambda:self.YouPress.emit(int(7)))
        self.pushButton_8.clicked.connect(lambda:self.YouPress.emit(int(8)))


class Monitor(QThread):
    WantToPrint=pyqtSignal(bytes)
    def __init__(self, parent=None):
        super().__init__(parent)
        self.text=[]
    def run(self):
        while 1:
            self.text=ser.read(20)
            self.WantToPrint.emit(self.text)
            self.text =[]

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.setupUi(self)
        self.now_angle_data=[3500,3500,4500,4000,6700,4000,6000,4900,4200,3500,4900,0,7500,4100,4000,4000,6800]
        self.accumulate_angle_data=[]
        self.num_of_active=0
        self.setting()
        self.vision_effect()
        self.control_object()
        self.total_table_set()
        self.RC_control=RemoteControl()
        self.RC_control.YouPress.connect(self.SimulateRemoteControl)
        self.dia_waiting=Waiting()
        self.monitor_win=Monitor()
        self.monitor_win.WantToPrint.connect(self.mmmmonitor)
        self.monitor_win.start()
        self.spinBox_11.setEnabled(False) #紅色4號殘障了QQ
        self.horizontalSlider_11.setEnabled(False) #紅色4號殘障了QQ

    def mmmmonitor(self,value): #印出M128端printf的東西 & 接收ACK
        for x in value:
            if x==255:
                self.dia_waiting.accept()  #將叫使用者不要亂動的對話窗關起來
                QMessageBox.about(self,"Done","播放完成")
        self.textBrowser.append(str(value))

    def setting(self):     #通訊埠設定
        items = listports()
        item,ok=QInputDialog.getItem(self,"選擇序列埠","Select Your COM",items,0,False)
        if ok:
            ser.port = str(item)
            ser.open()
            time.sleep(5)
        else:
            self.accept()

    def what_is_type(self,data):   #幫通訊封包定義的type
        return 1

    def encoder_and_send_pac(self,data):  #通訊封包
        tmp=data[1]<<8
        tmp=tmp+data[2]
        print('data=',data[0],tmp,data[3])
        pac=[]
        pac +=  [253]  #Header=0xfd
        typeeeee= self.what_is_type(data)
        pac+=bytes([typeeeee]) #type
        l=bytes([len(data)>>8,len(data)&0xff])
        pac+= l  #bytes(2bytes)
        pac += data  #data
        checksum= bytes([(sum(data)+sum(l))&0xFF])
        pac += checksum #checksum
        for x in pac:
            ser.write(bytes([x]))
            time.sleep(0.05)


    def total_table_set(self):  #總表設定
        self.tableWidget.setHorizontalHeaderLabels(['定格1','定格2','定格3','定格4','定格5','定格6','定格7','定格8','定格9','定格10'])
        self.tableWidget.setVerticalHeaderLabels(['ID_0','ID_1','ID_2','ID_3','ID_4','ID_5','ID_6','ID_7','ID_8','ID_9','ID_10','ID_11','ID_12','ID_13','ID_14','ID_15','ID_16','間格'])
        self.tableWidget.resizeColumnsToContents()  #調整格子寬度 與顯示內容配合
        self.tableWidget.resizeRowsToContents()  #調整格子長度 與顯示內容配合

    def total_table_update(self):  #總表更新
        self.lcdNumber.display(self.num_of_active) #LCD顯示值更新
        for x in range(180):
            if x < len(self.accumulate_angle_data):
                itemmm=QTableWidgetItem(str(self.accumulate_angle_data[x]))
                itemmm.setTextAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
                self.tableWidget.setItem(x%18,int(x/18),itemmm)
            else:
                itemmm=QTableWidgetItem(' ')
                itemmm.setTextAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
                self.tableWidget.setItem(x%18,int(x/18),itemmm)





    def vision_effect(self):   #設置一些圖片而已
        self.setWindowTitle('KHR-3HV')
        self.setWindowIcon(QIcon(':/image/icon.ico'))
        self.label.setPixmap(QPixmap(":/image/_background.jpg"))

    def control_object(self):     #按鈕們
        self.pushButton.clicked.connect(self.closefile)
        self.pushButton_2.clicked.connect(self.del_all)
        self.pushButton_3.clicked.connect(self.make_excel)
        self.pushButton_7.clicked.connect(lambda:self.RC_control.show())
        self.pushButton_8.clicked.connect(self.note_now_angle)
        self.pushButton_9.clicked.connect(self.to_the_best_position)
        self.spinBox.valueChanged['int'].connect(self.ID_0)
        self.spinBox_1.valueChanged['int'].connect(self.ID_1)
        self.spinBox_2.valueChanged['int'].connect(self.ID_2)
        self.spinBox_3.valueChanged['int'].connect(self.ID_3)
        self.spinBox_4.valueChanged['int'].connect(self.ID_4)
        self.spinBox_5.valueChanged['int'].connect(self.ID_5)
        self.spinBox_6.valueChanged['int'].connect(self.ID_6)
        self.spinBox_7.valueChanged['int'].connect(self.ID_7)
        self.spinBox_8.valueChanged['int'].connect(self.ID_8)
        self.spinBox_9.valueChanged['int'].connect(self.ID_9)
        self.spinBox_10.valueChanged['int'].connect(self.ID_10)
        # self.spinBox_11.valueChanged['int'].connect(self.ID_11)  #紅色4號殘障了QQ
        self.spinBox_12.valueChanged['int'].connect(self.ID_12)
        self.spinBox_13.valueChanged['int'].connect(self.ID_13)
        self.spinBox_14.valueChanged['int'].connect(self.ID_14)
        self.spinBox_15.valueChanged['int'].connect(self.ID_15)
        self.spinBox_16.valueChanged['int'].connect(self.ID_16)


    def SimulateRemoteControl(self,value):   #遙控器送訊的地方
        self.dia_waiting.show()
        self.encoder_and_send_pac([0,0,0,20+value])



    def ID_0(self):
        self.encoder_and_send_pac([0,self.spinBox.value()>>8,self.spinBox.value()&255,0])
        self.now_angle_data[0]=self.spinBox.value()
    def ID_1(self):
        self.encoder_and_send_pac([1,self.spinBox_1.value()>>8,self.spinBox_1.value()&255,0])
        self.now_angle_data[1]=self.spinBox_1.value()
    def ID_2(self):
        self.encoder_and_send_pac([2,self.spinBox_2.value()>>8,self.spinBox_2.value()&255,0])
        self.now_angle_data[2]=self.spinBox_2.value()
    def ID_3(self):
        self.encoder_and_send_pac([3,self.spinBox_3.value()>>8,self.spinBox_3.value()&255,0])
        self.now_angle_data[3]=self.spinBox_3.value()
    def ID_4(self):
        self.encoder_and_send_pac([4,self.spinBox_4.value()>>8,self.spinBox_4.value()&255,0])
        self.now_angle_data[4]=self.spinBox_4.value()
    def ID_5(self):
        self.encoder_and_send_pac([5,self.spinBox_5.value()>>8,self.spinBox_5.value()&255,0])
        self.now_angle_data[5]=self.spinBox_5.value()
    def ID_6(self):
        self.encoder_and_send_pac([6,self.spinBox_6.value()>>8,self.spinBox_6.value()&255,0])
        self.now_angle_data[6]=self.spinBox_6.value()
    def ID_7(self):
        self.encoder_and_send_pac([7,self.spinBox_7.value()>>8,self.spinBox_7.value()&255,0])
        self.now_angle_data[7]=self.spinBox_7.value()
    def ID_8(self):
        self.encoder_and_send_pac([8,self.spinBox_8.value()>>8,self.spinBox_8.value()&255,0])
        self.now_angle_data[8]=self.spinBox_8.value()
    def ID_9(self):
        self.encoder_and_send_pac([9,self.spinBox_9.value()>>8,self.spinBox_9.value()&255,0])
        self.now_angle_data[9]=self.spinBox_9.value()
    def ID_10(self):
        self.encoder_and_send_pac([10,self.spinBox_10.value()>>8,self.spinBox_10.value()&255,0])
        self.now_angle_data[10]=self.spinBox_10.value()
    # def ID_11(self):     #紅色4號殘障了QQ
    #     self.encoder_and_send_pac([11,self.spinBox_11.value()>>8,self.spinBox_11.value()&255,0])
    #     self.now_angle_data[11]=self.spinBox_11.value()
    def ID_12(self):
        self.encoder_and_send_pac([12,self.spinBox_12.value()>>8,self.spinBox_12.value()&255,0])
        self.now_angle_data[12]=self.spinBox_12.value()
    def ID_13(self):
        self.encoder_and_send_pac([13,self.spinBox_13.value()>>8,self.spinBox_13.value()&255,0])
        self.now_angle_data[13]=self.spinBox_13.value()
    def ID_14(self):
        self.encoder_and_send_pac([14,self.spinBox_14.value()>>8,self.spinBox_14.value()&255,0])
        self.now_angle_data[14]=self.spinBox_14.value()
    def ID_15(self):
        self.encoder_and_send_pac([15,self.spinBox_15.value()>>8,self.spinBox_15.value()&255,0])
        self.now_angle_data[15]=self.spinBox_15.value()
    def ID_16(self):
        self.encoder_and_send_pac([16,self.spinBox_16.value()>>8,self.spinBox_16.value()&255,0])
        self.now_angle_data[16]=self.spinBox_16.value()


    def to_the_best_position(self):  #強迫轉到最佳位置
        self.now_angle_data=[7500,8600,9500,4700,7300,7500,9000,8300,7500,6400,5500,0,7700,7500,6000,6700,7500]
        data=[0,0,0,31]
        self.encoder_and_send_pac(data)
        self.spinBox.setValue(7500)
        self.spinBox_1.setValue(8600)
        self.spinBox_2.setValue(9500)
        self.spinBox_3.setValue(4700)
        self.spinBox_4.setValue(7300)
        self.spinBox_5.setValue(7500)
        self.spinBox_6.setValue(9000)
        self.spinBox_7.setValue(8300)
        self.spinBox_8.setValue(7500)
        self.spinBox_9.setValue(6400)
        self.spinBox_10.setValue(5500)
        # self.spinBox_11.setValue(10300)    #紅色4號殘障了QQ
        self.spinBox_12.setValue(7700)
        self.spinBox_13.setValue(7500)
        self.spinBox_14.setValue(6000)
        self.spinBox_15.setValue(6700)
        self.spinBox_16.setValue(7500)

    def note_now_angle(self):  #更新accumulate_angle_data
        if self.num_of_active >=10:
            QMessageBox.about(self,"Too long","動作串長度不可以超過10個")
        else:
            if self.num_of_active ==0:
                items=("1","2","3","4","5","6","7","8")
                item,ok=QInputDialog.getItem(self,"SDC open file","選擇編號1~8",items,0,False)
                if ok and item:
                    data=[0,0,0,10+int(item)]
                    # self.dia_waiting.show()
                    self.encoder_and_send_pac(data)
                    self.accumulate_angle_data += self.now_angle_data
                    self.num_of_active +=1
                    self.total_table_update()  #更新總表

            elif self.num_of_active <10 and self.num_of_active>0:
                items=("1","2","3","4","5","6","7","8","9","10")
                item,ok=QInputDialog.getItem(self,"紀錄姿態","與前一姿態間格數",items,0,False)
                if ok:
                    self.accumulate_angle_data += [int(item)]
                    self.encoder_and_send_pac([0,0,0, int(item)])
                    self.accumulate_angle_data += self.now_angle_data
                    self.num_of_active +=1
                    self.total_table_update()  #更新總表


    def del_all(self):  #清空accumulate_angle_data
        if self.num_of_active >0:
            self.accumulate_angle_data = []
            self.num_of_active =0
            self.total_table_update()  #更新總表
            QMessageBox.about(self,"Clear","動作串已清空")
        else:
            QMessageBox.about(self,"No Thing","本來就沒有暫存之動作串")


    def closefile(self):   #CloseFile
        if self.accumulate_angle_data !=[]:
            data=[0,0,0,19]        #command=19
            # self.dia_waiting.show()
            self.encoder_and_send_pac(data)
        else:
            QMessageBox.about(self,"No Thing","尚未開啟SDC檔案")


    def make_excel(self):
        options = QFileDialog.Options()
        fileName, _=QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","(*.xlsx)", options=options)
        if fileName:
            print('filename=',fileName)
            wb = Workbook()# 創建一個工作薄
            sheet = wb.active# 創建一個工作表(注意是一個屬性)
            sheet.title = 'create_sheet'# excel創建的工作表名默認為sheet1,一下代碼實現了給新創建的工作表創建一個新的名字
            sheet['A19'] = '間格'
            for i in range(17):
                sheet.cell(row=i+2,column=1,value='ID %d'%i)
            for j in range(18):
                sheet.cell(row=1,column=j+2,value='定格 %d'%(j+1))
            self.accumulate_angle_data +=['#']
            for i in range(18):
                for j in range(self.num_of_active):
                    sheet.cell(row=i+2,column=j+2,value=self.accumulate_angle_data[18*j+i])
            self.accumulate_angle_data.pop()
            wb.save('create_excel.xlsx')# 保存一個文檔
            shutil.move("create_excel.xlsx",fileName)




if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
